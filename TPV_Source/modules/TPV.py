import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox as mBox
from pathlib import Path
import os.path
from datetime import datetime, date
import webbrowser
import logging

from configobj import ConfigObj
import openpyxl as op
from docx import Document

from modules import maintanance


class TPV_Main():

    def __init__(self, master, tabcontroll, name, config_name):
        
        self.config_name = config_name
        self.date = datetime.today()

        frametxt = 'TPV Skjema'

        if os.path.isfile(self.config_name) is True:
            self.config = ConfigObj(self.config_name, encoding='utf8', default_encoding='utf8')
            frametxt = self.config['Diversje']['5']

        self.master = master
        self.tab = tabcontroll

        self.tabvar = ttk.Frame(self.tab)

        self.tab.add(self.tabvar, text=name)
        self.tab.pack(expand=1, fill="both")

        # Instantiating a labelframe to contain the application in
        self.TPV_Body = ttk.LabelFrame(self.tabvar, text=frametxt)
        self.TPV_Body.pack(expand=1)

        menu_bar = tk.Menu(self.master)
        self.master.config(menu=menu_bar)

        file_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu = tk.Menu(menu_bar, tearoff=0)

        help_menu.add_command(label='Prosedyre', command=self.procedure)
        help_menu.add_command(label='Hjelp', command=self._op_wiki)
        menu_bar.add_cascade(label='Info', menu=help_menu)

        file_menu.add_command(label='Lagre Utført vedlikehold', command=self.maintanance)
        file_menu.add_command(label='Åpne eksisterende excel fil', command=self._op_saved)
        menu_bar.add_cascade(label='Alternativer', menu=file_menu)

        self._config_gen()
        self._movable_dates()
        self.main()

    def _config_gen(self):

        if not os.path.isfile(self.config_name):

            self.config = ConfigObj(encoding='utf8', default_encoding='utf8')
            self.config.filename = self.config_name

            self.config['Vedlikeholdspunkt'] = {}
            self.config['Vedlikeholdspunkt']['1'] = 'Put your info here'
            self.config['Vedlikeholdspunkt']['2'] = 'Put your info here'
            self.config['Vedlikeholdspunkt']['3'] = 'Put your info here'
            self.config['Vedlikeholdspunkt']['4'] = 'Put your info here'
            self.config['Vedlikeholdspunkt']['5'] = 'Put your info here'

            self.config['Handling'] = {}
            self.config['Handling']['1'] = 'Put your info here'
            self.config['Handling']['2'] = 'Put your info here'
            self.config['Handling']['3'] = 'Put your info here'
            self.config['Handling']['4'] = 'Put your info here'
            self.config['Handling']['5'] = 'Put your info here'

            self.config['Oljetype'] = {}
            self.config['Oljetype']['1'] = 'Put your info here'
            self.config['Oljetype']['2'] = 'Put your info here'
            self.config['Oljetype']['3'] = 'Put your info here'
            self.config['Oljetype']['4'] = 'Put your info here'
            self.config['Oljetype']['5'] = 'Put your info here'

            self.config['Hyppighet'] = {}
            self.config['Hyppighet']['1'] = 'Put your info here'
            self.config['Hyppighet']['2'] = 'Put your info here'
            self.config['Hyppighet']['3'] = 'Put your info here'
            self.config['Hyppighet']['4'] = 'Put your info here'
            self.config['Hyppighet']['5'] = 'Put your info here'

            self.config['Tidspunkt'] = {}
            self.config['Tidspunkt']['1'] = 'Put your info here'
            self.config['Tidspunkt']['2'] = 'Put your info here'
            self.config['Tidspunkt']['3'] = 'Put your info here'
            self.config['Tidspunkt']['4'] = 'Put your info here'
            self.config['Tidspunkt']['5'] = 'Put your info here'

            self.config['Intervall'] = {}
            self.config['Intervall']['1'] = '0'
            self.config['Intervall']['2'] = '0'
            self.config['Intervall']['3'] = '0'

            self.config['Dato'] = {}
            self.config['Dato']['1'] = '0'
            self.config['Dato']['2'] = '0'
            self.config['Dato']['3'] = '0'

            self.config['Diversje'] = {}
            self.config['Diversje']['1'] = 'Du kan skrive ekstra info her:'
            self.config['Diversje']['2'] = 'Annet:'
            self.config['Diversje']['3'] = self.date.strftime('%Y')
            self.config['Diversje']['4'] = '0'
            self.config['Diversje']['5'] = 'TPV Skjema for Maskin...'
            self.config['Diversje']['6'] = '0'

            self.config['Filbehandling'] = {}
            self.config['Filbehandling']['1'] = ''
            self.config['Filbehandling']['2'] = '850x450'
            self.config['Filbehandling']['3'] = ''
            self.config['Filbehandling']['4'] = ''

            self.config['Ukentlig'] = {}
            self.config['Ukentlig']['1'] = '0'

            self.config['Månedlig'] = {}
            self.config['Månedlig']['1'] = '0'

            self.config['Kvartalsvis'] = {}
            self.config['Kvartalsvis']['1'] = '0'

            self.config['Halvår'] = {}
            self.config['Halvår']['1'] = '0'

            self.config['Årlig'] = {}
            self.config['Årlig']['1'] = '0'

            self.config.write()

    def main(self):
        
        """Main body of the gui application"""

        # Accessing the config parser and getting the number of keys
        
        keys = self.config.keys()

        first_key = keys[0]

        # Collecting current date and weekday in full name
        today = self.date.strftime('%A')

        # Collecting which number in the month the current day is
        today_number = int(self.date.strftime('%d'))

        # Collecting which number in the year the current day is
        day_of_year = int(self.date.strftime('%j'))

        # Establish the number of entries in the config file
        values = [i for i in self.config[first_key]]

        length = len(values)
        
        day = self.day_check()

        key_weekly = 1
        self.weekly = {}

        key_monthly = 1
        self.monthly = {}

        key_quarterly = 1
        self.quarterly = {}

        key_halfyear = 1
        self.halfyear = {}

        key_yearly = 1
        self.yearly = {}

        intervalls_counter = 0

        row_ved = 1
        row_han = 1
        row_olj = 1
        row_hyp = 1
        row_when = 1

        # Generating labels on the fly based on how many entries it is in the config file
        for value in self.config['Vedlikeholdspunkt'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_ved, column=1, sticky=tk.W, padx=15)
            row_ved += 1

        for value in self.config['Handling'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_han, column=2, sticky=tk.W, padx=15)
            row_han += 1

        for value in self.config['Oljetype'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_olj, column=3, sticky=tk.W, padx=15)
            row_olj += 1

        for value in self.config['Hyppighet'].values():

            lowcas = value.lower()

            if lowcas == 'daglig':
                label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='green')
                label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                row_hyp += 1

            elif lowcas == 'ukentlig':

                weekly_flag = self.config['Ukentlig'][str(key_weekly)]

                if today == 'Friday' or weekly_flag == '1':

                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='yellow')
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.weekly['Weekly' + str(key_weekly)] = intervalls_counter
                    key_weekly += 1

                else:
                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.weekly['Weekly' + str(key_weekly)] = intervalls_counter
                    key_weekly += 1

            elif lowcas == 'månedlig':

                monthly_flag = self.config['Månedlig'][str(key_monthly)]

                if today_number == day or monthly_flag == '1':

                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='orange')
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.monthly['Monthly' + str(key_monthly)] = intervalls_counter
                    key_monthly += 1

                else:
                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.monthly['Monthly' + str(key_monthly)] = intervalls_counter
                    key_monthly += 1

            elif lowcas == 'kvartalsvis':

                quarterly_flag = self.config['Kvartalsvis'][str(key_quarterly)]

                if day_of_year == self.config['Dato']['1'] or quarterly_flag == '1':

                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='orange')
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.quarterly['Quarterly' + str(key_quarterly)] = intervalls_counter
                    key_quarterly += 1

                else:
                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.quarterly['Quarterly' + str(key_quarterly)] = intervalls_counter
                    key_quarterly += 1

            elif lowcas == 'halvår':

                halfyearly_flag = self.config['Halvår'][str(key_halfyear)]

                if day_of_year == self.config['Dato']['2'] or halfyearly_flag == '1':

                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='red')
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.halfyear['Halfyear' + str(key_halfyear)] = intervalls_counter
                    key_halfyear += 1

                else:
                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.halfyear['Halfyear' + str(key_halfyear)] = intervalls_counter
                    key_halfyear += 1

            elif lowcas == 'årlig':

                yearly_flag = self.config['Årlig'][str(key_yearly)]

                if day_of_year == self.config['Dato']['3'] or yearly_flag == '1':

                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='red')
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.yearly['Yearly' + str(key_yearly)] = intervalls_counter
                    key_yearly += 1

                else:
                    label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                    label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                    row_hyp += 1

                    self.yearly['Yearly' + str(key_yearly)] = intervalls_counter
                    key_yearly += 1

            else:
                label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                label.grid(row=row_hyp, column=4, sticky=tk.W, padx=15)
                row_hyp += 1

            intervalls_counter += 1

        for value in self.config['Tidspunkt'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_when, column=5, sticky=tk.N, padx=15)
            row_when += 1

        self.results = {}
        check_boxes = {}

        for i in range(length):
            self.results['checkvar{0}'.format(i)] = tk.IntVar()
        
        row = 1
        for item in self.results:
            check_boxes['check{0}'.format(i)] = ttk.Checkbutton(self.TPV_Body, variable=self.results[item]).grid(row=row, column=0, sticky=tk.W)
            row += 1

        lbl1 = tk.Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
        lbl1.grid(row=0, column=1, sticky=tk.N)
        lbl2 = tk.Label(self.TPV_Body, text='Handling:', font=FONT2)
        lbl2.grid(row=0, column=2, sticky=tk.N)
        lbl3 = tk.Label(self.TPV_Body, text='Oljetype:', font=FONT2)
        lbl3.grid(row=0, column=3, sticky=tk.N)
        lbl4 = tk.Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
        lbl4.grid(row=0, column=4, sticky=tk.N)
        lbl5 = tk.Label(self.TPV_Body, text='Tidspunkt:', font=FONT2)
        lbl5.grid(row=0, column=5, sticky=tk.N)

        header = self.config['Diversje']['1']
        lbl16 = tk.Label(self.TPV_Body, text=header)
        lbl16.grid(row=row, column=0, columnspan=2,  pady=3)
        row += 1

        self.txt = tk.Text(self.TPV_Body, height=3, width=25)
        self.txt.grid(row=row, column=0, columnspan=2)
        row += 1

        button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
        button.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=15)

    def save(self):
        

        values = [self.results[i] for i in self.results]

        self.checkbox_values = [i.get() for i in values]

        # Main code for writing out the excel file used to save the data in comes here:

        # Calling the config parser and accessing filepath if present in config file
        f_name = self.config['Filbehandling']['1']

        # Dealing with dates related to where in the excel file stuff will be saved
        month = str(self.date.strftime('%B'))
        day_as_string = str(int(self.date.strftime('%d')) + 1)
        day_as_int = int(self.date.strftime('%d')) + 1
        today = self.date.strftime('%d.%m.%Y, %a')

        cell = 'A' + day_as_string

        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                  'August', 'September', 'October', 'November', 'December']

        index = [i for i in self.config['Vedlikeholdspunkt'].values()]

        index.append(self.config['Diversje']['2'])


        # Checking if the file exist or if we have switched year
        if not os.path.isfile(f_name) or self._year_check() == 1:

            self._error_popup('Saving', 'Make sure you arent overwriting an old file')

            f_new = filedialog.asksaveasfilename(title='Select File',
                                                 filetypes=(("Excel files", ".xlsx"),
                                                            ("All files", "*.*")), 
                                                 defaultextension="*.*")
            
            self.config['Filbehandling']['1'] = f_new

            wb = op.Workbook()

            for i in months:
                wb.create_sheet(i)

            del wb['Sheet']
            wb_sheets = wb.sheetnames

            col = 2

            for wb_month in wb_sheets:
                ws = wb[wb_month]
                for i in index:
                    ws.cell(row=1, column=col, value=i)
                    col += 1
                    if col > len(index) + 1:
                        col = 2
                    else:
                        continue

            ws = wb[month]

            ws[cell] = today

            col = 2

            for i in self.checkbox_values:
                ws.cell(row=day_as_int, column=col, value=i)
                col += 1

            try:
                op.writer.excel.save_workbook(wb, f_new)
                mBox.showinfo('', 'Resultater har blitt lagret')
                self.config.write()

            except FileNotFoundError as e:
                self._error_popup('Error!', e)

        elif os.path.isfile(f_name) and self._year_check() == 0:

            f_name = self.config['Filbehandling']['1']

            wb = op.load_workbook(filename=f_name)
            ws = wb[month]

            ws[cell] = today

            col = 2

            for i in self.checkbox_values:
                ws.cell(row=day_as_int, column=col, value=i)
                col += 1

            textbox = self.txt.get('1.0', tk.END)
            ws.cell(row=day_as_int, column=col, value=textbox)

            self._persistent_weekly(self.checkbox_values, wb)

            op.writer.excel.save_workbook(wb, f_name)

            mBox.showinfo('', 'Resultater har blitt lagret')

        else:
            self._error_popup('Major Error!', 'The whole universe has shattered, take cover!')

    def _year_check(self):

        """A simple method for checking if we have switched year"""

        config_year = self.config['Diversje']['3']

        current_year = self.date.strftime('%Y')

        if config_year != current_year:
            self.config['Diversje']['3'] = current_year
            self.config.write()
            return 1
        else:
            return 0

    def _op_saved(self):

        """Lets you open a preexisting excel file"""

        f_exist = filedialog.askopenfilename()

        self.config['Filbehandling']['1'] = f_exist
        self.config.write()

    def _op_wiki(self):

        """Simply opens your default browser and directs you to the wiki"""

        webbrowser.open('https://github.com/UniQueKakarot/TPV_Skjema/wiki')

    def procedure(self):

        """Lets you select a word file that gets linked to in the UI"""

        f_exist = self.config['Filbehandling']['3']

        if os.path.isfile(f_exist) is True:
            os.system(f_exist)

        elif os.path.isfile(f_exist) is False:
            f_exist = filedialog.askopenfilename()
            os.system(f_exist)
            self.config['Filbehandling']['3'] = f_exist
            self.config.write()

        else:
            self._error_popup("Procedure failed", "We could not locate or save the procedure document")

    def maintanance(self):
        
        """Adding in a new window for saving extra information on maintenance done"""

        maintanance.Maintanace(self.config, self.date, FONT1, FONT2)

    def day_check(self, year=None, month=None, day=20):

        """Checks to see if the 20th day of the month falls on a weekend"""

        if year == None:
            year = int(self.date.strftime('%Y'))

        if month == None:
            month = int(self.date.strftime('%m'))
        
        check_day = date(year, month, day).isoweekday()

        if check_day == 6:
            day -= 1
            return day
        
        elif check_day == 7:
            day += 1
            return day

        else:
            return 20

    def _persistent_weekly(self, checkbox_values, workbook):

        """ Monitoring the maintainance so that correct maintainance is done on time
            and if not signal that to the user by lighting up the labels """

        today = self.date.strftime('%A')
        #today = 'Thursday'
        today_number = int(self.date.strftime('%d'))
        month_number = int(self.date.strftime('%m'))
        day_of_year = int(self.date.strftime('%j'))

        backwards_counting = {'Saturday': 1, 'Sunday': 2, 'Monday': 3, 'Tuesday': 4, 'Wednesday': 5, 'Thursday': 6, 'Friday': 0}
        check_pos = today_number - backwards_counting[today]

        # I hate working with dates!
        worksheet_months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 
                            9: 'September', 10: 'October', 11: 'November', 12: 'December'}

        days_in_month = {'January': 31, 'February': 28, 'March': 31, 'April': 30, 'May': 31, 'June': 30, 'July': 31, 'August': 31,
                         'September': 30, 'October': 31, 'November': 30, 'December': 31}

        # if friday is so close to the start of the month that when we check back
        # we need to check the previous month, subtract 1 of month number and convert to
        # month name and set the correct worksheet and get the right number of days to subtract

        try:
            if check_pos < 1:
                worksheet = workbook[worksheet_months[(month_number - 1)]]
                check_pos = days_in_month[worksheet_months[(month_number - 1)]] - abs(check_pos)

            else:
                worksheet = workbook[worksheet_months[month_number]]
        except KeyError:
            print('if check_pos < 1 failed somewhere')
            pass

        # weekly
        config_pos = 1
        if today == 'Friday':
            
            for i in self.weekly.values():

                if checkbox_values[i] == 1:
                    self.config['Ukentlig'][str(config_pos)] = '0'
                    config_pos += 1
                else:
                    self.config['Ukentlig'][str(config_pos)] = '1'
                    config_pos += 1

        else:

            for position in self.weekly.values():
                position += 2

                try:
                    if worksheet.cell(row=check_pos, column=position).value == 0 or worksheet.cell(row=check_pos, column=position).value == None:
                        self.config['Ukentlig'][str(config_pos)] = '1'
                except NameError:
                    print('worksheet not defined')
                    pass
                
                config_pos += 1

        config_pos = 1
        for value, location in zip(self.config['Ukentlig'].values(), self.weekly.values()):

            if value == '1' and checkbox_values[location] == 1:
                self.config['Ukentlig'][str(config_pos)] = '0'
                location += 2
                
                try:
                    worksheet.cell(row=check_pos, column=location, value=1)
                except NameError:
                    print('worksheet not defined 2')
                    pass

            config_pos += 1


        # monthly
        config_pos = 1
        if today_number == self.day_check():
            
            for i in self.monthly.values():

                if checkbox_values[i] == 1:
                    self.config['Månedlig'][str(config_pos)] = '0'
                    config_pos += 1
                else:
                    self.config['Månedlig'][str(config_pos)] = '1'
                    config_pos += 1

        else:

            try:
                worksheet = workbook[worksheet_months[(month_number - 1)]]
                prev_month = month_number - 1
            
                for position in self.monthly.values():
                    position += 2

                    if worksheet.cell(row=self.day_check(month=prev_month), column=position).value == 0 or worksheet.cell(row=self.day_check(month=prev_month), column=position).value == None:
                        self.config['Månedlig'][str(config_pos)] = '1'
                    
                    config_pos += 1
            except KeyError:
                print('I failed you master... q.q')
                pass

        config_pos = 1
        for value, location in zip(self.config['Månedlig'].values(), self.monthly.values()):
            
            if value == '1' and checkbox_values[location] == 1:
                self.config['Månedlig'][str(config_pos)] = '0'

                try:
                    worksheet = workbook[worksheet_months[(month_number - 1)]]
                    prev_month = month_number - 1
                    location += 2
                    worksheet.cell(row=self.day_check(month=prev_month), column=location, value=1)
                except KeyError:
                    print('Shit hit the fan')
                    pass
        
            config_pos += 1

        # quarterly
        config_pos = 1
        if day_of_year == self.config['Dato']['1']:

            for i in self.quarterly.values():

                if checkbox_values[i] == 1:
                    self.config['Kvartalsvis'][str(config_pos)] = '0'
                    config_pos += 1
                else:
                    self.config['Kvartalsvis'][str(config_pos)] = '1'
                    config_pos += 1

        config_pos = 1
        for value, location in zip(self.config['Kvartalsvis'].values(), self.quarterly.values()):
            
            if value == '1' and checkbox_values[location] == 1:
                self.config['Kvartalsvis'][str(config_pos)] = '0'

            config_pos += 1

        # halfyearly
        config_pos = 1
        if day_of_year == self.config['Dato']['2']:

            for i in self.halfyear.values():

                if checkbox_values[i] == 1:
                    self.config['Halvår'][str(config_pos)] = '0'
                    config_pos += 1
                else:
                    self.config['Halvår'][str(config_pos)] = '1'
                    config_pos += 1

        config_pos = 1
        for value, location in zip(self.config['Halvår'].values(), self.halfyear.values()):
            
            if value == '1' and checkbox_values[location] == 1:
                self.config['Halvår'][str(config_pos)] = '0'

            config_pos += 1

        # yearly
        config_pos = 1
        if day_of_year == self.config['Dato']['3']:

            for i in self.yearly.values():

                if checkbox_values[i] == 1:
                    self.config['Årlig'][str(config_pos)] = '0'
                    config_pos += 1
                else:
                    self.config['Årlig'][str(config_pos)] = '1'
                    config_pos += 1

        config_pos = 1
        for value, location in zip(self.config['Årlig'].values(), self.yearly.values()):
            
            if value == '1' and checkbox_values[location] == 1:
                self.config['Årlig'][str(config_pos)] = '0'

            config_pos += 1

        self.config.write()

    def _movable_dates(self):

        """ Deal with incremental maintainance times instead of a fixed date """

        for i in self.config['Intervall']:

            intervall = int(self.config['Intervall'][str(i)])
            old_intervall = int(self.config['Dato'][str(i)])
            day_of_year = int(self.date.strftime('%j'))

            if old_intervall <= day_of_year:
                new_intervall = old_intervall + intervall

                if new_intervall > 366:
                    new_intervall %= 366
                    

                self.config['Dato'][str(i)] = str(new_intervall)

        self.config.write()

    def _error_popup(self, message, issue):

        mBox.showerror('{}'.format(message), '{}'.format(issue))


#win = tk.Tk()
#win.title("TPV Skjema")
#win.geometry("850x460")

#config_name = 'config.ini'
#config_file = Path('TPV-Skjema/config.ini')

FONT1 = ("Calibri", 11)
FONT2 = ("Calibri", 11, "bold", "underline")

#tpv = TPV_Main()

#win.mainloop()
