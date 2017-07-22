import openpyxl as op
from datetime import datetime, date

#wb = op.Workbook() #Opens up a new excel workbook
wb = op.load_workbook(filename = 'test1.xlsx')

date = datetime.today()

month = date.strftime('%B')
month = str(month)

day = date.strftime('%d')
day = str(day)

today = date.strftime('%d.%m.%Y, %a')

#Sheet names:
months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
          'August', 'September', 'October', 'November', 'December']

colum_head = ['Spindel', 'Spindel Oljetank', 'Hydraulisk Enhet', 'Kjølevannsystem',
              'Maskindeksling & ATC', 'Roterende Akser', 'Kabinett Filter', 'Verktøymagasin', 'Spontransportør']

dataset1 = [1, 1, 1, 0, 0, 0, 0, 0, 0]
dataset2 = [0, 0, 0, 1, 1, 1, 0, 0, 0]
dataset3 = [0, 0, 0, 0, 0, 0, 1, 1, 1]

for i in months:
    wb.create_sheet(i)

#sh = wb.get_sheet_names()
#x = wb.get_sheet_by_name('Sheet')
#wb.remove_sheet(x)

ws = wb[month]

col = 2

for i in colum_head:
    ws.cell(row=1, column=col, value=i)
    col += 1

cell = 'A' + day
print(cell)

ws[cell] = today



col = 2

for i in dataset3:
    ws.cell(row=24, column=col, value=i)
    col += 1

wb.save('test1.xlsx')

# for sheet in wb:
#     print(sheet.title)
