# coding: cp865

from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox as mBox
from pathlib import Path
from configobj import ConfigObj
import os.path
from datetime import datetime, date, timedelta
import webbrowser
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows

win = Tk()
win.title("TPV Skjema")
win.geometry("850x460")

config_name = 'config.ini'
config_file = Path('TPV-Skjema/config.ini')

FONT1 = ("Calibri", 11)
FONT2 = ("Calibri", 11, "bold", "underline")


class TPV_Main():

    def __init__(self):

        self.TPV_Body = ttk.LabelFrame(win, text='TPV Skjema')
        self.TPV_Body.pack(expand=1)

        menuBar = Menu(win)
        win.config(menu=menuBar)

        fileMenu = Menu(menuBar, tearoff=0)
        helpMenu = Menu(menuBar, tearoff=0)

        helpMenu.add_command(label='Hjelp', command=self.op_wiki)
        menuBar.add_cascade(label='Info', menu=helpMenu)

        #fileMenu.add_command(label='Lagre ekstra info')
        fileMenu.add_command(label='Aapne eksisterende excel fil', command=self.op_saved)
        menuBar.add_cascade(label='Alternativer', menu=fileMenu)


        self.config()
        self.main()

    def config(self):

        if os.path.isfile('config.ini') == True:
            config1 = ConfigObj('config.ini')

        else:
            config = ConfigObj()
            config.filename = config_name

            config['Vedlikeholdspunkt'] = {}
            config['Vedlikeholdspunkt']['1'] = 'Put your info here'
            config['Vedlikeholdspunkt']['2'] = 'Put your info here'
            config['Vedlikeholdspunkt']['3'] = 'Put your info here'
            config['Vedlikeholdspunkt']['4'] = 'Put your info here'
            config['Vedlikeholdspunkt']['5'] = 'Put your info here'
            config['Vedlikeholdspunkt']['6'] = 'Put your info here'
            config['Vedlikeholdspunkt']['7'] = 'Put your info here'
            config['Vedlikeholdspunkt']['8'] = 'Put your info here'
            config['Vedlikeholdspunkt']['9'] = 'Put your info here'

            config['Handling'] = {}
            config['Handling']['1'] = 'Put your info here'
            config['Handling']['2'] = 'Put your info here'
            config['Handling']['3'] = 'Put your info here'
            config['Handling']['4'] = 'Put your info here'
            config['Handling']['5'] = 'Put your info here'
            config['Handling']['6'] = 'Put your info here'
            config['Handling']['7'] = 'Put your info here'
            config['Handling']['8'] = 'Put your info here'
            config['Handling']['9'] = 'Put your info here'

            config['Oljetype'] = {}
            config['Oljetype']['1'] = 'Put your info here'
            config['Oljetype']['2'] = 'Put your info here'
            config['Oljetype']['3'] = 'Put your info here'
            config['Oljetype']['4'] = 'Put your info here'
            config['Oljetype']['5'] = 'Put your info here'
            config['Oljetype']['6'] = 'Put your info here'
            config['Oljetype']['7'] = 'Put your info here'
            config['Oljetype']['8'] = 'Put your info here'
            config['Oljetype']['9'] = 'Put your info here'

            config['Hyppighet'] = {}
            config['Hyppighet']['1'] = 'Put your info here'
            config['Hyppighet']['2'] = 'Put your info here'
            config['Hyppighet']['3'] = 'Put your info here'
            config['Hyppighet']['4'] = 'Put your info here'
            config['Hyppighet']['5'] = 'Put your info here'
            config['Hyppighet']['6'] = 'Put your info here'
            config['Hyppighet']['7'] = 'Put your info here'
            config['Hyppighet']['8'] = 'Put your info here'
            config['Hyppighet']['9'] = 'Put your info here'

            config['Diversje'] = {}
            #config['Diversje']['1'] = 'Du kan skrive ekstra info her:'

            config['Filbehandling'] = {}
            config['Filbehandling']['1'] = '' #Filename stored here
            config['Filbehandling']['2'] = '' #Filepath stored here

            config.write()



    def main(self):

        config1 = ConfigObj('config.ini') #Config Parser
        keys = config1.keys() #Getting the number of keys in the config
        values = [] #Empty list for storing how many entries it is in each key
        first_key = keys[0] #First key in the config for entries list
        date = datetime.today() #Getting the current date
        today = date.strftime('%A') #Getting the current day, as in day name. Mon, tue, wed and such
        today_number = date.strftime('%d')
        today_number = int(today_number)
        day_of_year = date.strftime('%j')
        day_of_year = int(day_of_year)

        #Establish the number of entries in the config file
        for i in config1[first_key]:
            values.append(i)

        global length
        length = len(values)


        row_ved = 1
        row_han = 1
        row_olj = 1
        row_hyp = 1

        for value in config1['Vedlikeholdspunkt'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_ved, column=1, sticky=W, padx=15)
            row_ved += 1


        for value in config1['Handling'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_han, column=2, sticky=W, padx=15)
            row_han += 1

        for value in config1['Oljetype'].values():
            label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
            label.grid(row=row_olj, column=3, sticky=W, padx=15)
            row_olj += 1


        for value in config1['Hyppighet'].values():
            lowCas = value.lower() #converting the string value in value to all lower case for safety

            if lowCas == 'daglig':
                label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='green')
                label.grid(row=row_hyp, column=4, sticky=W, padx=15)
                row_hyp += 1

            elif lowCas == 'ukentlig' and today == 'Friday':

                label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='yellow')
                label.grid(row=row_hyp, column=4, sticky=W, padx=15)
                row_hyp += 1

            elif today_number == 20 and lowCas == 'maanedlig':

                label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='orange')
                label.grid(row=row_hyp, column=4, sticky=W, padx=15)
                row_hyp += 1

            elif lowCas == 'halvaar' and day_of_year == 183:

                label = ttk.Label(self.TPV_Body, text=value, font=FONT1, background='red')
                label.grid(row=row_hyp, column=4, sticky=W, padx=15)
                row_hyp += 1

            else:
                label = ttk.Label(self.TPV_Body, text=value, font=FONT1)
                label.grid(row=row_hyp, column=4, sticky=W, padx=15)
                row_hyp += 1


        if length == '5':
            print('5')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()

            self.list5 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5]

            check1 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt')
            lbl1.grid(row=0, column=1, sticky=N)
            lbl2 = Label(self.TPV_Body, text='Handling')
            lbl2.grid(row=0, column=2, sticky=N)
            lbl3 = Label(self.TPV_Body, text='Oljetype')
            lbl3.grid(row=0, column=3, sticky=N)
            lbl4 = Label(self.TPV_Body, text='Hyppighet')
            lbl4.grid(row=0, column=4, sticky=N)

            button = ttk.Button(self.TPV_Body, text='Lagre')
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 6:
            print('6')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()

            self.list6 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6]

            check1 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt')
            lbl1.grid(row=0, column=1, sticky=N)
            lbl2 = Label(self.TPV_Body, text='Handling')
            lbl2.grid(row=0, column=2, sticky=N)
            lbl3 = Label(self.TPV_Body, text='Oljetype')
            lbl3.grid(row=0, column=3, sticky=N)
            lbl4 = Label(self.TPV_Body, text='Hyppighet')
            lbl4.grid(row=0, column=4, sticky=N)

            button = ttk.Button(self.TPV_Body, text='Lagre')
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 7:
            print('7')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()

            self.list7 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7]

            check1 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, text=label, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt')
            lbl1.grid(row=0, column=1, sticky=N)
            lbl2 = Label(self.TPV_Body, text='Handling')
            lbl2.grid(row=0, column=2, sticky=N)
            lbl3 = Label(self.TPV_Body, text='Oljetype')
            lbl3.grid(row=0, column=3, sticky=N)
            lbl4 = Label(self.TPV_Body, text='Hyppighet')
            lbl4.grid(row=0, column=4, sticky=N)

            button = ttk.Button(self.TPV_Body, text='Lagre')
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 8:
            print('8')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()

            self.list8 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt')
            lbl1.grid(row=0, column=1, sticky=N)
            lbl2 = Label(self.TPV_Body, text='Handling')
            lbl2.grid(row=0, column=2, sticky=N)
            lbl3 = Label(self.TPV_Body, text='Oljetype')
            lbl3.grid(row=0, column=3, sticky=N)
            lbl4 = Label(self.TPV_Body, text='Hyppighet')
            lbl4.grid(row=0, column=4, sticky=N)

            button = ttk.Button(self.TPV_Body, text='Lagre')
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 9:
            #print('something = 9')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()

            self.list9 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)

            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.Save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 10:
            print('10')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()
            self.checkVar10 = IntVar()

            self.list10 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9,
                           self.checkVar10]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)
            check10 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar10)
            check10.grid(row=10, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)

            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 11:
            print('11')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()
            self.checkVar10 = IntVar()
            self.checkVar11 = IntVar()

            self.list11 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9,
                          self.checkVar10, self.checkVar11]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)
            check10 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar10)
            check10.grid(row=10, column=0, sticky=W)
            check11 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar11)
            check11.grid(row=11, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)

            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 12:
            print('12')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()
            self.checkVar10 = IntVar()
            self.checkVar11 = IntVar()
            self.checkVar12 = IntVar()

            self.list12 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9,
                          self.checkVar10, self.checkVar11, self.checkVar12]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)
            check10 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar10)
            check10.grid(row=10, column=0, sticky=W)
            check11 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar11)
            check11.grid(row=11, column=0, sticky=W)
            check12 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar12)
            check12.grid(row=12, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)

            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 13:
            print('13')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()
            self.checkVar10 = IntVar()
            self.checkVar11 = IntVar()
            self.checkVar12 = IntVar()
            self.checkVar13 = IntVar()

            self.list13 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9,
                          self.checkVar10, self.checkVar11, self.checkVar12, self.checkVar13]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)
            check10 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar10)
            check10.grid(row=10, column=0, sticky=W)
            check11 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar11)
            check11.grid(row=11, column=0, sticky=W)
            check12 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar12)
            check12.grid(row=12, column=0, sticky=W)
            check13 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar13)
            check13.grid(row=13, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)

            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 14:
            print('14')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()
            self.checkVar10 = IntVar()
            self.checkVar11 = IntVar()
            self.checkVar12 = IntVar()
            self.checkVar13 = IntVar()
            self.checkVar14 = IntVar()

            self.list14 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9,
                          self.checkVar10, self.checkVar11, self.checkVar12, self.checkVar13, self.checkVar14]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)
            check10 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar10)
            check10.grid(row=10, column=0, sticky=W)
            check11 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar11)
            check11.grid(row=11, column=0, sticky=W)
            check12 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar12)
            check12.grid(row=12, column=0, sticky=W)
            check13 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar13)
            check13.grid(row=13, column=0, sticky=W)
            check14 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar14)
            check14.grid(row=14, column=0, sticky=W)

            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)

            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        elif length == 15:
            print('15')
            self.checkVar1 = IntVar()
            self.checkVar2 = IntVar()
            self.checkVar3 = IntVar()
            self.checkVar4 = IntVar()
            self.checkVar5 = IntVar()
            self.checkVar6 = IntVar()
            self.checkVar7 = IntVar()
            self.checkVar8 = IntVar()
            self.checkVar9 = IntVar()
            self.checkVar10 = IntVar()
            self.checkVar11 = IntVar()
            self.checkVar12 = IntVar()
            self.checkVar13 = IntVar()
            self.checkVar14 = IntVar()
            self.checkVar15 = IntVar()

            self.list15 = [self.checkVar1, self.checkVar2, self.checkVar3, self.checkVar4, self.checkVar5, self.checkVar6, self.checkVar7, self.checkVar8, self.checkVar9,
                          self.checkVar10, self.checkVar11, self.checkVar12, self.checkVar13, self.checkVar14, self.checkVar15]

            check1 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar1)
            check1.grid(row=1, column=0, sticky=W)
            check2 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar2)
            check2.grid(row=2, column=0, sticky=W)
            check3 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar3)
            check3.grid(row=3, column=0, sticky=W)
            check4 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar4)
            check4.grid(row=4, column=0, sticky=W)
            check5 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar5)
            check5.grid(row=5, column=0, sticky=W)
            check6 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar6)
            check6.grid(row=6, column=0, sticky=W)
            check7 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar7)
            check7.grid(row=7, column=0, sticky=W)
            check8 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar8)
            check8.grid(row=8, column=0, sticky=W)
            check9 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar9)
            check9.grid(row=9, column=0, sticky=W)
            check10 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar10)
            check10.grid(row=10, column=0, sticky=W)
            check11 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar11)
            check11.grid(row=11, column=0, sticky=W)
            check12 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar12)
            check12.grid(row=12, column=0, sticky=W)
            check13 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar13)
            check13.grid(row=13, column=0, sticky=W)
            check14 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar14)
            check14.grid(row=14, column=0, sticky=W)
            check15 = ttk.Checkbutton(self.TPV_Body, variable=self.checkVar15)
            check15.grid(row=15, column=0, sticky=W)


            lbl1 = Label(self.TPV_Body, text='Vedlikeholdspunkt:', font=FONT2)
            lbl1.grid(row=0, column=1, sticky=N, pady=5)
            lbl2 = Label(self.TPV_Body, text='Handling:', font=FONT2)
            lbl2.grid(row=0, column=2, sticky=N, pady=5)
            lbl3 = Label(self.TPV_Body, text='Oljetype:', font=FONT2)
            lbl3.grid(row=0, column=3, sticky=W, pady=5)
            lbl4 = Label(self.TPV_Body, text='Hyppighet:', font=FONT2)
            lbl4.grid(row=0, column=4, sticky=N, pady=5)


            button = ttk.Button(self.TPV_Body, text='Lagre', command=self.save)
            button.grid(row=21, column=0, columnspan=2, sticky=W, pady=15)

        else:
            print('initializing failed! Values out of bounds')


    def Save(self):

        values1 = []
        values2 = []

        if length == 5:
            for i in self.list5:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 6:
            for i in self.list6:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 7:
            for i in self.list7:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 8:
            for i in self.list8:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 9:
            for i in self.list9:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 10:
            for i in self.list10:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 11:
            for i in self.list11:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 12:
            for i in self.list12:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 13:
            for i in self.list13:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 14:
            for i in self.list14:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        elif length == 15:
            for i in self.list15:
                values1.append(i)

            for i in values1:
                x = i.get()
                values2.append(x)

        else:
            print('Well looks like shit hit the fan son!')



        #stuff concerning the creation of a dataframe comes under here

        config = ConfigObj('config.ini', encoding='utf8') #Config Parser
        f_name = config['Filbehandling']['1']             #reading in the location and name for the file

        date = datetime.today()                           #getting info on the current day

        month = date.strftime('%B')
        month = str(month)

        day_as_string = date.strftime('%d')
        day_as_string = str(day_as_string)

        day_as_int = date.strftime('%d')
        day_as_int = int(day_as_int)

        today = date.strftime('%d.%m.%Y, %a')             #formatting the date info to my liking

        cell = 'A' + day_as_string

        index = []

        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
        'August', 'September', 'October', 'November', 'December']



        for i in config['Vedlikeholdspunkt'].values():
            index.append(i)



        #checking if the file exist or not
        if os.path.isfile(f_name) == False:

            f_new = filedialog.asksaveasfilename(title='Select File') #ask's you where to save the file

            config['Filbehandling']['1'] = f_new                      #assigning the filename a place in the config file
            config.write()                                            #writing the filename to the config file

            wb = op.Workbook()

            for i in months:
                wb.create_sheet(i)

            sh = wb.get_sheet_names()
            x = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(x)

            ws = wb[month]

            col = 2

            for i in config1['Vedlikeholdspunkt'].values():
                ws.cell(row=1, column=col, value=i)
                col += 1

            ws[cell] = today

            col = 2

            for i in values2:

            op.writer.excel.save_workbook(wb, f_new)

            df[today] = pd.Series(values2, index=df.index)         #generating a seires off the results

            mBox.showinfo('', 'Resultater har blitt lagret')       #showing the user a visual feedback when the result are saved


        elif os.path.isfile(f_name) == True:

            f_name = config['Filbehandling']['1']                     #reading in the location and name for the file

            wb = op.load_workbook(filename=f_name)
            ws = wb[month]

            # for i in dataframe_to_rows(df, index=True, header=True):
            #     ws.append(i)
            #
            # op.writer.excel.save_workbook(wb, f_name)

            mBox.showinfo('', 'Resultater har blitt lagret')          #showing the user a visual feedback when the result are saved

        else:
            print('Looks like something went wrong!')


    def op_saved(self):

        config = ConfigObj('config.ini', encoding='utf8')             #Config Parser

        f_exist = filedialog.askopenfilename()                        #open a window to let the user select the already saved file

        config['Filbehandling']['1'] = f_exist                        #assigning the value a place in the config file
        config.write()                                                #writing the filename to the config file

    def op_wiki(self):

        wiki = webbrowser.open('https://github.com/UniQueKakarot/TPV_Skjema/wiki')




tpv = TPV_Main()

win.mainloop()
